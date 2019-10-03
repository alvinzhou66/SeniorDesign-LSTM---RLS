import xlrd
import openpyxl
import numpy as np

# Handling sentiment analyzed results
# calculate 7 functions and form new excel document

fileName = r'D:\课程\Senior Project\DATA\Weibo\sentiment analysis\重庆\cq.xlsx'

wb = xlrd.open_workbook(fileName)
sheet1 = wb.sheet_by_name('Sheet1')
rowNum = sheet1.nrows
# 把两列数据放在一个列表中 put two columns into list_data[]
data_list = []
data_time = []
data_value = []
for i in range(rowNum):
    if sheet1.cell(i, 0).ctype == 3:

        date = xlrd.xldate.xldate_as_datetime(sheet1.cell(i, 0).value, 0)
        dateStr = str(date.year)+u'/'+str(date.month)+u'/'+str(date.day)
        # print(dateStr)
        data_time.append(dateStr)

    else:
        data_time.append(sheet1.cell(i, 0).value)
    i += 1
data_value.extend(sheet1.col_values(1))
counter = 1
while counter < len(data_time):
    data_list.append([data_time[counter], data_value[counter]])
    # print(data_list[counter - 1])
    counter += 1

print(len(data_list))

# group by day
data = {}  # the dictionary saving the data group by day
for x in range(len(data_list)):
    if data_list[x][0] in data.keys():
        data[data_list[x][0]].append(data_list[x][1])
    else:
        data[data_list[x][0]] = [data_list[x][1]]

# traverse the dictionary and calculate
# 7 functions: min, max, mean, standard deviation, range, kurtosis, skewness
data_excel = []  # saving the data will put into the excel
for key, value in data.items():
    R = np.array(value)
    dataMin = min(value)
    dataMax = max(value)
    dataMean = np.mean(value)
    dataStd = np.std(value)
    dataRange = dataMax - dataMin
    if dataStd == 0:
        continue
    dataKu = np.mean((R - dataMean) ** 4) / pow(dataStd, 4)
    dataSc = np.mean(((R - dataMean) ** 3) / pow(dataStd, 3))
    data_excel.append([key, dataMin, dataMax, dataMean, dataStd, dataRange, dataKu, dataSc])

print(len(data))

# write the data into a new excel
title = ['time', 'min', 'max', 'mean', 'standard deviation', 'range', 'kurtosis', 'skewness']
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.append(title)
for row in data_excel:
    ws1.append(row)
newFileName = fileName[0:len(fileName)-5] + "_summary" + fileName[len(fileName)-5:len(fileName)]
wb.save(newFileName)


