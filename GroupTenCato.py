import xlrd
import openpyxl
import numpy as np

# Handling sentiment analyzed results
# calculate 7 functions and form new excel document

# AQI + date from real data set
realFileName = r'D:\课程\Senior Project\DATA\new_air_data_for_the_ten_cities\重庆 201502-201902.xlsx'
wbr = xlrd.open_workbook(realFileName)
sheet = wbr.sheet_by_name('Sheet1')
rowNum2 = sheet.nrows
dataTime = []
dataAQI = []
for ii in range(rowNum2):
    if sheet.cell(ii, 0).ctype == 3:

        date = xlrd.xldate.xldate_as_datetime(sheet.cell(ii, 0).value, 0)
        dateStr = str(date.year)+u'/'+str(date.month)+u'/'+str(date.day)
        # print(dateStr)
        dataTime.append(dateStr)

    else:
        dataTime.append(sheet.cell(ii, 0).value)
    ii += 1
dataAQI.extend(sheet.col_values(2))
dataCombine = []
count = 1
while count < len(dataTime):
    dataCombine.append([dataTime[count], dataAQI[count]])
    count += 1
print(len(dataCombine))
# print(dataCombine)
fileName = r'D:\课程\Senior Project\DATA\Weibo\sentiment analysis\重庆\cq.xlsx'

wb = xlrd.open_workbook(fileName)
sheet1 = wb.sheet_by_name('Sheet1')
rowNum = sheet1.nrows
# 把两列数据放在一个列表中 put two columns into list_data[]
data_list = []
data_time = []
data_value = []
for i in range(rowNum):
    if sheet1.cell(i, 0).ctype == 3:

        date = xlrd.xldate.xldate_as_datetime(sheet1.cell(i, 0).value, 0)
        dateStr = str(date.year)+u'/'+str(date.month)+u'/'+str(date.day)
        # print(dateStr)
        data_time.append(dateStr)

    else:
        data_time.append(sheet1.cell(i, 0).value)
    i += 1
data_value.extend(sheet1.col_values(1))

counter = 1
while counter < len(data_time):
    data_list.append([data_time[counter], data_value[counter]])
    # print(data_list[counter - 1])
    counter += 1

print(len(data_list))

# group by day
data = {}  # the dictionary saving the data group by day
for x in range(len(data_list)):
    if data_list[x][0] in data.keys():
        data[data_list[x][0]].append(data_list[x][1])
    else:
        data[data_list[x][0]] = [data_list[x][1]]

# traverse the dictionary and calculate
# 10-dimension and 4-dimension
data_excel = []  # saving the data will put into the excel
for key, value in data.items():
    a1, a2, a3, a4, a5, a6, a7, a8, a9, a10 = 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
    R = np.array(value)
    dataMin = min(value)
    dataMax = max(value)
    dataMean = np.mean(value)
    dataStd = np.std(value)
    dataRange = dataMax - dataMin
    if dataStd == 0:
        continue
    dataKu = np.mean((R - dataMean) ** 4) / pow(dataStd, 4)
    dataSc = np.mean(((R - dataMean) ** 3) / pow(dataStd, 3))
    for y in range(len(dataCombine)):
        innerList = []
        innerList.extend(dataCombine[y])
        # print(innerList)
        if key in innerList:
            # print(key)
            # print(dataCombine.index(innerList))
            list_values = []
            list_values.extend(value)

            for xx in range(len(list_values)):
                if 0.0 <= list_values[xx] < 0.25:
                    a1 += 1
                elif 0.25 <= list_values[xx] < 0.5:
                    a2 += 1
                elif 0.5 <= list_values[xx] < 0.75:
                    a3 += 1
                elif 0.75 <= list_values[xx] < 1.0:
                    a4 += 1
            data_excel.append([key, a1/len(list_values), a2/len(list_values), a3/len(list_values),
                               a4/len(list_values), dataMin, dataMax, dataMean, dataStd, dataRange, dataKu, dataSc,
                               dataCombine[dataCombine.index(innerList)][1]])
            #     if 0.0 <= list_values[xx] < 0.1:
            #         a1 += 1
            #     elif 0.1 <= list_values[xx] < 0.2:
            #         a2 += 1
            #     elif 0.2 <= list_values[xx] < 0.3:
            #         a3 += 1
            #     elif 0.3 <= list_values[xx] < 0.4:
            #         a4 += 1
            #     elif 0.4 <= list_values[xx] < 0.5:
            #         a5 += 1
            #     elif 0.5 <= list_values[xx] < 0.6:
            #         a6 += 1
            #     elif 0.6 <= list_values[xx] < 0.7:
            #         a7 += 1
            #     elif 0.7 <= list_values[xx] < 0.8:
            #         a8 += 1
            #     elif 0.8 <= list_values[xx] < 0.9:
            #         a9 += 1
            #     elif 0.9 <= list_values[xx] < 1.0:
            #         a10 += 1
            # data_excel.append([key, a1/len(list_values), a2/len(list_values), a3/len(list_values),
            #                    a4/len(list_values),
            #                    a5/len(list_values), a6/len(list_values), a7/len(list_values), a8/len(list_values),
            #                    a9/len(list_values),  a10/len(list_values),
            #                    dataCombine[dataCombine.index(innerList)][1]])
        else:
            continue
print(len(data))
print(len(data_excel))

# write the data into a new excel
title = ['time', 'a1', 'a2', 'a3', 'a4', 'min', 'max', 'mean', 'standard deviation', 'range', 'kurtosis',
         'skewness', 'AQI']
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.append(title)
for row in data_excel:
    ws1.append(row)
newFileName = fileName[0:len(fileName)-5] + "_s3" + fileName[len(fileName)-5:len(fileName)]
wb.save(newFileName)
