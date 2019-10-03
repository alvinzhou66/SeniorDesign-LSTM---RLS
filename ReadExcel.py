# -*- encoding: utf-8 -*-
from __future__ import print_function, unicode_literals
import json
import requests
import xlrd
import openpyxl


# 读取EXCEL表 read sheet of excel document
# 对其中的评论或博文那一列，进行语义分析 do sentiment analysis on the column with comments or posts
# 将日期和语义分析结果写入一个新的EXCEL表，该表自动生成在源文件所在目录并添加后缀“_senti”
# the date and results of sentiment analysis are written into a new excel, the file end with "_senti"

fileName = r'D:\课程\Senior Project\DATA\Weibo\深圳\深圳天气微博.xlsx'

SENTIMENT_URL = 'http://api.bosonnlp.com/sentiment/analysis'
# Changing X-Token as user
headers = {
    'X-Token': '8sFM66ob.34061.X0YBbv8pj0AY',
    'Content-Type': 'application/json'
}

# 索引表，需要依据具体文件  sheet name depends on the actual sheet
wb = xlrd.open_workbook(fileName)
sheet1 = wb.sheet_by_name('Sheet1')
rowNum = sheet1.nrows
x = rowNum / 100
# print(x)

# 依据具体的表，修改需要操作的列索引 according to the table, fix the index of columns
data_list = []
data_list.extend(sheet1.col_values(2))
# for item in data_list:
#     print(item)
data_res = []  # 存放循环中语义分析结果的所有结果 save all the results from sentiment analysis in the While loop
n = 0
while n < x:
    # print(data_list[(n*10+1):((n+1)*10+1)])
    data = json.dumps(data_list[(n*100+1):((n+1)*100+1)])
    resp = requests.post(SENTIMENT_URL, headers=headers, data=data.encode('utf-8'))
    print(resp.text)  # 每次出来的结果都是text (the results are in text)
    data_res.extend(eval(resp.text))
    n += 1

# 将相关数据写入新的excel表，文件名后缀为‘_senti’  data written in a new excel , file name end with "_senti"
newFileName = fileName[0:len(fileName)-5] + "_senti" + fileName[len(fileName)-5:len(fileName)]

# 日期, 若读取的列为日期格式，将其转化为字符串
# date information saved in data_time[], if read Date format, transfer to String
data_time = []
for i in range(rowNum):
    if sheet1.cell(i, 0).ctype == 3:

        date = xlrd.xldate.xldate_as_datetime(sheet1.cell(i, 0).value, 0)
        dateStr = str(date.year)+u'/'+str(date.month)+u'/'+str(date.day)
        print(dateStr)
        data_time.append(dateStr)

    else:
        data_time.append(sheet1.cell(i, 0).value)
    i += 1

data = []  # 新列表--新建excel中的内容  saving all the data in new excel

# 把日期和语义分析结果放入新的列表中 saving process
counter = 0
while counter < len(data_res):
    data.append([data_time[counter + 1], data_res[counter][0], data_res[counter][1]])
    print(data[counter])
    counter += 1

title = ['time', 'positive', 'negative']
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.append(title)
for row in data:
    ws1.append(row)

wb.save(newFileName)








